"""
Excel Export module for DPLUS Dashboard.
Exports dashboard data to formatted Excel files.
"""

import pandas as pd
from io import BytesIO
from datetime import date, datetime
from typing import Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.logger import log_error, log_info


# Style definitions
HEADER_FILL = PatternFill(start_color="4A7C6F", end_color="4A7C6F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)
NUMBER_FORMAT = '#,##0'
CURRENCY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.0%'
BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

# Segment colors
HERO_FILL = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
CORE_FILL = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
VOLUME_FILL = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")


def export_to_excel(
    data: Dict[str, pd.DataFrame],
    filters: Dict,
    metrics: Optional[Dict] = None
) -> bytes:
    """
    Export dashboard data to a formatted Excel file.

    Args:
        data: Dictionary of DataFrames (revenue, aov, products, etc.)
        filters: Filter settings used
        metrics: Summary metrics

    Returns:
        Excel file as bytes
    """
    try:
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create Summary sheet
        _create_summary_sheet(wb, filters, metrics)

        # Create Revenue sheet
        if 'revenue' in data and not data['revenue'].empty:
            _create_revenue_sheet(wb, data['revenue'])

        # Create AOV sheet
        if 'aov' in data and not data['aov'].empty:
            _create_aov_sheet(wb, data['aov'])

        # Create Products sheet
        if 'products' in data and not data['products'].empty:
            _create_products_sheet(wb, data['products'])

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        log_info("Excel export generated successfully", {
            'sheets': list(data.keys()),
            'start_date': str(filters.get('start_date')),
            'end_date': str(filters.get('end_date'))
        })

        return output.getvalue()

    except Exception as e:
        log_error(e, {'operation': 'export_to_excel'})
        raise


def _create_summary_sheet(wb: Workbook, filters: Dict, metrics: Optional[Dict]):
    """Create the summary sheet with KPIs."""
    ws = wb.create_sheet("Summary")

    # Title
    ws['A1'] = "D Plus Skin Analytics Report"
    ws['A1'].font = Font(bold=True, size=18, color="4A7C6F")
    ws.merge_cells('A1:E1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10, color="64748B")

    # Date range
    ws['A4'] = "Report Period"
    ws['A4'].font = SUBHEADER_FONT
    ws['A5'] = f"From: {filters.get('start_date', 'N/A')}"
    ws['A6'] = f"To: {filters.get('end_date', 'N/A')}"
    ws['A7'] = f"Platform: {filters.get('platform', 'All')}"
    ws['A8'] = f"Granularity: {filters.get('granularity', 'Daily')}"

    # Metrics section
    if metrics:
        ws['A10'] = "Key Performance Indicators"
        ws['A10'].font = Font(bold=True, size=14, color="4A7C6F")
        ws.merge_cells('A10:C10')

        row = 12
        metric_labels = {
            'total_revenue': ('Total Revenue', CURRENCY_FORMAT),
            'total_orders': ('Total Orders', NUMBER_FORMAT),
            'total_quantity': ('Total Quantity', NUMBER_FORMAT),
            'aov': ('Average Order Value', CURRENCY_FORMAT),
            'unique_products': ('Unique Products', NUMBER_FORMAT)
        }

        for key, (label, fmt) in metric_labels.items():
            if key in metrics:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = metrics[key]
                ws[f'B{row}'].number_format = fmt
                row += 1

    # Segment legend
    ws[f'A{row + 2}'] = "Segment Legend"
    ws[f'A{row + 2}'].font = Font(bold=True, size=12)

    ws[f'A{row + 4}'] = "Hero (Max)"
    ws[f'A{row + 4}'].fill = HERO_FILL
    ws[f'A{row + 4}'].font = Font(color="FFFFFF")
    ws[f'B{row + 4}'] = "Top 20% highest revenue days"

    ws[f'A{row + 5}'] = "Core (Middle)"
    ws[f'A{row + 5}'].fill = CORE_FILL
    ws[f'A{row + 5}'].font = Font(color="FFFFFF")
    ws[f'B{row + 5}'] = "Average performance"

    ws[f'A{row + 6}'] = "Volume (Min)"
    ws[f'A{row + 6}'].fill = VOLUME_FILL
    ws[f'A{row + 6}'].font = Font(color="FFFFFF")
    ws[f'B{row + 6}'] = "Bottom 20% lowest revenue days"

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15


def _create_revenue_sheet(wb: Workbook, df: pd.DataFrame):
    """Create revenue data sheet."""
    ws = wb.create_sheet("Revenue")

    # Header
    ws['A1'] = "Revenue Trends"
    ws['A1'].font = Font(bold=True, size=14, color="4A7C6F")

    # Data table
    _write_dataframe(ws, df, start_row=3)

    # Column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15


def _create_aov_sheet(wb: Workbook, df: pd.DataFrame):
    """Create AOV data sheet."""
    ws = wb.create_sheet("AOV Analysis")

    # Header
    ws['A1'] = "Average Order Value Analysis"
    ws['A1'].font = Font(bold=True, size=14, color="4A7C6F")

    # Data table
    _write_dataframe(ws, df, start_row=3)

    # Column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15


def _create_products_sheet(wb: Workbook, df: pd.DataFrame):
    """Create products data sheet."""
    ws = wb.create_sheet("Products")

    # Header
    ws['A1'] = "Product Analysis"
    ws['A1'].font = Font(bold=True, size=14, color="4A7C6F")

    # Data table
    _write_dataframe(ws, df, start_row=3)

    # Column widths
    ws.column_dimensions['A'].width = 50  # Product name
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1):
    """Write DataFrame to worksheet with formatting."""
    # Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

    # Data rows
    for row_idx, row in enumerate(df.values, start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Apply formatting based on column
            col_name = df.columns[col_idx - 1]
            if 'revenue' in col_name.lower() or 'aov' in col_name.lower():
                cell.number_format = CURRENCY_FORMAT
            elif 'quantity' in col_name.lower() or 'orders' in col_name.lower():
                cell.number_format = NUMBER_FORMAT
            elif 'percentage' in col_name.lower() or 'pct' in col_name.lower():
                cell.number_format = PERCENT_FORMAT

            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')

            # Segment coloring
            if 'segment' in col_name.lower():
                if 'max' in str(value).lower() or 'hero' in str(value).lower():
                    cell.fill = HERO_FILL
                    cell.font = Font(color="FFFFFF")
                elif 'middle' in str(value).lower() or 'core' in str(value).lower():
                    cell.fill = CORE_FILL
                    cell.font = Font(color="FFFFFF")
                elif 'min' in str(value).lower() or 'volume' in str(value).lower():
                    cell.fill = VOLUME_FILL
                    cell.font = Font(color="FFFFFF")


def get_export_filename(filters: Dict, export_type: str = 'excel') -> str:
    """Generate a descriptive filename for the export."""
    start = filters.get('start_date', date.today())
    end = filters.get('end_date', date.today())
    platform = filters.get('platform', 'All')

    if isinstance(start, date):
        start = start.strftime('%Y%m%d')
    if isinstance(end, date):
        end = end.strftime('%Y%m%d')

    ext = 'xlsx' if export_type == 'excel' else 'pdf'

    return f"DPLUS_Report_{platform}_{start}_to_{end}.{ext}"
